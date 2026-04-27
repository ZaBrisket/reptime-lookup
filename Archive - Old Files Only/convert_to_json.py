"""Convert 'Who Makes the Best? Guide' Excel file to a queryable JSON database."""
import json
import re
from datetime import datetime
import openpyxl

SRC_PT1 = "Who-Makes-the-Best-Guide (Pt 1 of 2).xlsx"
SRC_PT2 = "Who-Makes-the-Best-Guide (Pt 2 of 2).xlsx"
DST = "app/who-makes-the-best-guide.json"

COLOR_TO_TIER = {
    "FFFFE599": "NWBIG",      # yellow → Best of the best
    "FFB6D7A8": "Super Rep",  # green  → Super Rep (pt 1 shade)
    "FFB7D7A8": "Super Rep",  # green  → Super Rep (pt 2 shade, 1-step variant)
}

# Section header rows (the row containing 'Model Family' / 'Movement' / ...)
# determined empirically — each section's first model family confirms the brand.
PT1_SECTIONS = [
    {"brand": "Rolex",                "header_row": 28},
    {"brand": "Patek Philippe",       "header_row": 122},
    {"brand": "Audemars Piguet",      "header_row": 190},
    {"brand": "Vacheron Constantin",  "header_row": 237},
    {"brand": "Richard Mille",        "header_row": 263},
    {"brand": "Omega",                "header_row": 295},
    {"brand": "Cartier",              "header_row": 334},
    {"brand": "Tudor",                "header_row": 374},
    {"brand": "Jaeger LeCoultre",     "header_row": 403},
    {"brand": "Panerai",              "header_row": 430},
    {"brand": "Breitling",            "header_row": 482},
    {"brand": "IWC",                  "header_row": 508},
]

# Pt 2 is a single legacy-reference section: "Old Panerai".
PT2_SECTIONS = [
    {"brand": "Panerai (Old / Legacy reference)", "header_row": 4, "legacy": True},
]

RANK_LABELS = ["Best Factory", "2nd Best", "3rd Best", "4th Best", "5th Best",
               "6th Best", "7th Best", "8th Best", "9th Best", "10th Best"]


def cell_color(cell):
    if cell.fill and cell.fill.fill_type == "solid" and cell.fill.fgColor:
        rgb = cell.fill.fgColor.rgb
        return str(rgb) if rgb is not None else None
    return None


def tier_for_color(rgb):
    return COLOR_TO_TIER.get(rgb)


def clean(v):
    """Normalize a cell value: trim strings, drop nbsp, leave numbers/dates alone."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace("\xa0", " ").strip()
        return v if v else None
    if isinstance(v, datetime):
        return v.date().isoformat()
    return v


def merged_value(ws, row, col, merged_lookup):
    """If (row, col) is inside a merged range, return the value of its anchor cell."""
    key = (row, col)
    if key in merged_lookup:
        ar, ac = merged_lookup[key]
        return clean(ws.cell(ar, ac).value)
    return clean(ws.cell(row, col).value)


def build_merged_lookup(ws):
    """Map every (row, col) inside a merged range → its anchor (top-left) cell."""
    lookup = {}
    for mr in ws.merged_cells.ranges:
        anchor = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                lookup[(r, c)] = anchor
    return lookup


def is_blank_row(ws, row, cols=range(1, 7)):
    return all(clean(ws.cell(row, c).value) is None for c in cols)


def looks_like_subsection(value):
    """Detect repeated 'HOW DO I ORDER...' callouts that interrupt brand sections."""
    if not isinstance(value, str):
        return False
    return "HOW DO I ORDER" in value.upper() or "click here" in value.lower()


def section_end(ws, start_row, next_section_start):
    """Find the last data row of a section before next_section_start."""
    end = next_section_start - 1 if next_section_start else ws.max_row
    while end > start_row and is_blank_row(ws, end):
        end -= 1
    return end


def parse_section(ws, merged_lookup, brand, header_row, end_row):
    """Walk rows below the header, emit one entry per data row.

    Each entry preserves: model family (resolved via merged cell),
    model number/notes, movement, ranked factory recommendations
    (with raw cell color → tier), and the source row number.
    """
    entries = []
    last_model_number = None  # carries forward when a row only adds a movement
    for r in range(header_row + 1, end_row + 1):
        # Skip the embedded "HOW DO I ORDER" callouts and any fully blank rows.
        first = clean(ws.cell(r, 1).value)
        second = clean(ws.cell(r, 2).value)
        if looks_like_subsection(first) or looks_like_subsection(second):
            continue
        if is_blank_row(ws, r, cols=range(1, 8)):
            continue

        family = merged_value(ws, r, 1, merged_lookup)
        model_number = merged_value(ws, r, 2, merged_lookup)
        movement = clean(ws.cell(r, 3).value)

        # Stand-alone notes (e.g. B50:F52) come through as one big string in col 2
        # without a movement or factory. Capture them as note entries.
        recs_raw = [(c, clean(ws.cell(r, c).value), cell_color(ws.cell(r, c)))
                    for c in range(4, 8)]
        recs = []
        for idx, (c, val, color) in enumerate(recs_raw):
            if val is None:
                continue
            recs.append({
                "rank": idx + 1,
                "rank_label": RANK_LABELS[idx] if idx < len(RANK_LABELS) else f"#{idx+1}",
                "factory": val,
                "tier": tier_for_color(color),
            })

        # If the row only carries a long note and nothing else useful, store it.
        if recs == [] and movement is None and model_number and len(str(model_number)) > 80:
            entries.append({
                "type": "note",
                "brand": brand,
                "model_family": family,
                "text": model_number,
            })
            continue

        # Inherit model_number from prior row when only a movement variant is added.
        if model_number is None and movement is not None and recs:
            model_number = last_model_number
        elif model_number is not None:
            last_model_number = model_number

        if family is None and model_number is None and movement is None and not recs:
            continue

        entries.append({
            "type": "watch",
            "brand": brand,
            "model_family": family,
            "model_number": model_number,
            "movement": movement,
            "recommendations": recs,
        })
    return entries


def parse_metadata(ws, merged_lookup):
    md = {
        "title": clean(ws.cell(1, 1).value),
        "order_link_text": clean(ws.cell(2, 1).value),
        "description": clean(ws.cell(4, 1).value),
        "how_to_order_url": "https://www.reddit.com/r/RepTime/wiki/index/",
    }
    # Updated date (B11) and update notes (B12)
    md["updated"] = clean(ws.cell(11, 2).value)
    md["update_notes"] = clean(ws.cell(12, 2).value)
    md["full_updates_link_text"] = clean(ws.cell(14, 2).value)

    # Brands covered (rows 15-18, cols B-D)
    brands = []
    for r in range(15, 19):
        for c in range(2, 5):
            v = clean(ws.cell(r, c).value)
            if v:
                brands.append(v)
    md["brands_covered"] = brands

    # Legend (rows 11-12, col E)
    md["legend"] = {
        "NWBIG": clean(ws.cell(11, 5).value),
        "Super Rep": clean(ws.cell(12, 5).value),
    }
    return md


def process_workbook(src_path, sections):
    """Parse one workbook and return (brands_out, flat_entries) lists."""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    merged_lookup = build_merged_lookup(ws)

    brands_out = []
    flat = []
    for i, sec in enumerate(sections):
        next_start = sections[i + 1]["header_row"] if i + 1 < len(sections) else None
        end_row = section_end(ws, sec["header_row"], next_start)
        entries = parse_section(ws, merged_lookup, sec["brand"], sec["header_row"], end_row)

        families = {}
        family_order = []
        notes = []
        for e in entries:
            if e["type"] == "note":
                notes.append({
                    "brand": e["brand"],
                    "model_family": e["model_family"],
                    "text": e["text"],
                })
                continue
            fam = e["model_family"] or "(unspecified)"
            if fam not in families:
                families[fam] = []
                family_order.append(fam)
            families[fam].append({
                "model_number": e["model_number"],
                "movement": e["movement"],
                "recommendations": e["recommendations"],
            })
            flat.append(e)

        brand_record = {
            "brand": sec["brand"],
            "model_families": [
                {"family": fam, "variants": families[fam]} for fam in family_order
            ],
            "notes": notes,
        }
        if sec.get("legacy"):
            brand_record["legacy"] = True
        brands_out.append(brand_record)
    return brands_out, flat, ws, merged_lookup


def main():
    pt1_brands, pt1_flat, pt1_ws, pt1_merged = process_workbook(SRC_PT1, PT1_SECTIONS)
    pt2_brands, pt2_flat, _, _ = process_workbook(SRC_PT2, PT2_SECTIONS)

    metadata = parse_metadata(pt1_ws, pt1_merged)

    brands_out = pt1_brands + pt2_brands
    flat = pt1_flat + pt2_flat

    # Build a compact factory index for "show me all watches where X is recommended"
    factory_index = {}
    for e in flat:
        for rec in e["recommendations"]:
            f = rec["factory"]
            factory_index.setdefault(f, []).append({
                "brand": e["brand"],
                "model_family": e["model_family"],
                "model_number": e["model_number"],
                "movement": e["movement"],
                "rank": rec["rank"],
                "tier": rec["tier"],
            })

    out = {
        "metadata": metadata,
        "brands": brands_out,
        "watches": [
            {k: v for k, v in e.items() if k != "type"} for e in flat
        ],
        "factory_index": factory_index,
        "stats": {
            "brand_count": len(brands_out),
            "watch_count": len(flat),
            "factory_count": len(factory_index),
            "model_family_count": sum(len(b["model_families"]) for b in brands_out),
        },
    }

    with open(DST, "w", encoding="utf-8") as fh:
        json.dump(out, fh, indent=2, ensure_ascii=False)

    print(f"Wrote {DST}")
    print(f"  brands: {out['stats']['brand_count']}")
    print(f"  model families: {out['stats']['model_family_count']}")
    print(f"  watch entries: {out['stats']['watch_count']}")
    print(f"  unique factories: {out['stats']['factory_count']}")


if __name__ == "__main__":
    main()
